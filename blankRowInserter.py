#! python3
# blankRowInserter.py - Starting at row N, inserts M blank rows into
# excel file, and returns it as a new excel file. (The original excel file
# remains unchanged).
# Usage:
#   ./blankRowInserter.py <N> <M> <file>

import openpyxl, sys

N = int(sys.argv[1])
M = int(sys.argv[2])
fileName = sys.argv[3]

wb = openpyxl.load_workbook(fileName)
ws = wb.active
newWb = openpyxl.Workbook()
newWs = newWb.active

for j in range(1, ws.max_column + 1):
    for i in range(1, ws.max_row + 1):
        if i < N:
            newWs.cell(row = i, column = j).value = \
                ws.cell(row = i, column = j).value
        else:
            newWs.cell(row = i + M, column = j).value = \
                ws.cell(row = i, column = j).value
        
newWb.save('new_' + fileName)


