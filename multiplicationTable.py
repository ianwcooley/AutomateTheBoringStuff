#! python3

# multiplicationTable.py - take a number N and makes an excel file of the
# N x N multiplication table
# Usage:
#   ./multiplicationTable.py <N>

import openpyxl, sys
from openpyxl.styles import Font

N = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, N+1):
    sheet.cell(row = 1, column = i+1).value = i
    sheet.cell(row = 1, column = i+1).font = Font(bold=True)
    sheet.cell(row = i+1, column = 1).value = i
    sheet.cell(row = i+1, column = 1).font = Font(bold=True)
    for j in range(1, N+1):
        sheet.cell(row = i + 1, column = j + 1).value = i * j
sheet.freeze_panes = 'B2'
wb.save(str(N) + 'x' + str(N) + '_multiplication_table.xlsx')
