#! python3

# excelToCsv.py - reads all the excel files in the cwd and outputs
# them as CSV files

import openpyxl, csv, os

# Loop over every excel file in the cwd
for excelFileName in os.listdir('.'):
    if not excelFileName.endswith('.xlsx'):
        continue
    print('writing ' + excelFileName)
    # Open the workbook for the excel file
    wb = openpyxl.load_workbook(excelFileName)
    # Loop through every sheet in the workbook.
    for sheetName in wb.sheetnames:
        ws = wb[sheetName]
        # make a csv file for the sheet
        csvFileName = os.path.splitext(excelFileName)[0] + sheetName + '.csv'
        # Create the csv.writer object for this CSV file.
        csvFile = open(csvFileName, 'w', newline='')
        csvWriter = csv.writer(csvFile)
        # Loop through every row in the sheet.
        for rowNum in range(1, ws.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, ws.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(str(ws.cell(row=rowNum, column=colNum).value))
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFile.close()